#!/usr/bin/env python3
"""
build_scaffold.py — esqueleto do workbook de valuation (skill valuation-consultivo).

Cria um .xlsx com as 9 abas da metodologia convergida, cabeçalhos, formatação
financeira e exemplos de fórmula viva. O agente preenche com os dados da conversa
(ou o usuário preenche os inputs azuis) e recalcula com o recalc.py da skill xlsx.

Uso:
    python build_scaffold.py [caminho_de_saida.xlsx]
"""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- Convenções (skill xlsx) ----
F_INPUT = Font(color="0000FF", name="Arial", size=11)      # inputs/hardcodes
F_FORM = Font(color="000000", name="Arial", size=11)       # fórmulas
F_LINK = Font(color="008000", name="Arial", size=11)       # link entre abas
F_HEAD = Font(bold=True, name="Arial", size=11)
F_TITLE = Font(bold=True, name="Arial", size=14)
FILL_HEAD = PatternFill("solid", fgColor="D9E1F2")
FILL_KEY = PatternFill("solid", fgColor="FFFF00")          # premissa-chave
FILL_SEC = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CUR = '"$"#,##0;("$"#,##0);-'
CUR_MM = '"$"#,##0.0,,"mm";("$"#,##0.0,,"mm");-'
PCT = "0.0%"
NUM = "#,##0"


def title(ws, text, span):
    ws["A1"] = text
    ws["A1"].font = F_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)


def header(ws, row, cols):
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = F_HEAD
        cell.fill = FILL_HEAD
        cell.border = BORDER


def box(ws, row, col, value, font=F_FORM, fill=None, num=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    if num:
        cell.number_format = num
    cell.border = BORDER
    return cell


def widths(ws, w):
    for i, x in enumerate(w, start=1):
        ws.column_dimensions[get_column_letter(i)].width = x


def build(path):
    wb = Workbook()
    wb.remove(wb.active)

    # ---------- 1. Premissas ----------
    ws = wb.create_sheet("Premissas")
    title(ws, "PREMISSAS — todos os inputs (azul = editar; amarelo = chave)", 7)
    header(ws, 3, ["Premissa", "Pessimista", "Provavel", "Otimista", "Fonte", "Confianca", "Status"])
    prem = [
        ("Custo de capital (r)", 0.20, 0.20, 0.20, "benchmark setor", "MEDIA", "validar"),
        ("Prob. sucesso Fase 1->2", 0.50, 0.66, 0.75, "Wong et al. 2019", "ALTA", "ok"),
        ("Prob. sucesso Fase 2->3", 0.45, 0.58, 0.70, "Wong et al. 2019", "ALTA", "ok"),
        ("Prob. aprovacao (pos-F3)", 0.45, 0.59, 0.70, "Wong et al. 2019", "ALTA", "ok"),
        ("Pico de receita (R$ mm)", 50, 100, 180, "premissa do usuario", "BAIXA", "a validar"),
        ("Anos ate o pico", 6, 8, 10, "premissa do usuario", "BAIXA", "a validar"),
        ("Anos de exclusividade", 8, 10, 12, "premissa do usuario", "BAIXA", "a validar"),
        ("Investimento da rodada (R$ mm)", 10, 10, 10, "usuario", "ALTA", "ok"),
        ("Option pool pos-rodada", 0.10, 0.10, 0.10, "usuario", "ALTA", "ok"),
    ]
    r = 4
    for p in prem:
        for j, v in enumerate(p, start=1):
            f = F_INPUT
            fill = FILL_KEY if p[0] in ("Custo de capital (r)", "Pico de receita (R$ mm)", "Investimento da rodada (R$ mm)") else None
            num = PCT if isinstance(v, float) else CUR_MM if "receita" in p[0] else NUM
            if j >= 5:
                f = F_FORM
                fill = None
                num = None
            box(ws, r, j, v, font=f, fill=fill, num=num)
        r += 1
    widths(ws, [34, 14, 12, 12, 30, 12, 14])

    # ---------- 2. Mercado ----------
    ws = wb.create_sheet("Mercado")
    title(ws, "MERCADO — TAM/SAM/SOM (bottom-up)", 6)
    header(ws, 3, ["Item", "Pessimista", "Provavel", "Otimista", "Fonte", "Obs"])
    mkt = [
        ("Populacao/area total (pacientes ou ha)", 100000, 500000, 1000000, "usuario", ""),
        ("Preco por unidade (R$)", 10, 15, 25, "usuario", ""),
        ("Penetracao no pico", 0.02, 0.05, 0.10, "usuario", ""),
        ("TAM (R$ mm)", "=B4*C4", None, None, "", "formula exemplo"),
    ]
    r = 4
    for m in mkt:
        for j, v in enumerate(m, start=1):
            num = PCT if isinstance(v, float) else CUR_MM if j in (2, 3, 4) else None
            box(ws, r, j, v, font=F_INPUT if j <= 4 else F_FORM, num=num)
        r += 1
    widths(ws, [42, 14, 12, 12, 30, 24])

    # ---------- 3. Modelo (rNPV) ----------
    ws = wb.create_sheet("Modelo")
    title(ws, "MODELO — rNPV (fluxos x prob. acumulada, descontados a r)", 14)
    header(ws, 3, ["Item"] + [f"Ano {i}" for i in range(0, 12)])
    rows = [
        "Transicao no ano (prob)",
        "Prob. acumulada de sucesso",
        "Receita (R$ mm)",
        "Custo (R$ mm)",
        "Fluxo liquido (R$ mm)",
        "Fluxo x prob. acumulada",
        "Fator de desconto",
        "Valor presente (R$ mm)",
    ]
    r = 4
    for name in rows:
        box(ws, r, 1, name, font=F_HEAD)
        r += 1
    # fórmulas de exemplo na linha 4..11, colunas B..M (ano 0..11)
    for col in range(2, 14):
        L = get_column_letter(col)
        y = col - 2
        box(ws, 4, col, 0.0, font=F_INPUT, num=PCT)                                   # transição (input)
        if col == 2:
            box(ws, 5, col, "=B4", font=F_FORM, num=PCT)                              # P acum ano0 = transição ano0
        else:
            box(ws, 5, col, f"=B5*{L}4", font=F_FORM, num=PCT)                        # produto acumulado
        box(ws, 6, col, None, font=F_INPUT, num=CUR_MM)                               # receita (input)
        box(ws, 7, col, None, font=F_INPUT, num=CUR_MM)                               # custo (input)
        box(ws, 8, col, f"={L}6-{L}7", font=F_FORM, num=CUR_MM)                       # fluxo líquido
        box(ws, 9, col, f"={L}8*{L}5", font=F_FORM, num=CUR_MM)                       # fluxo x prob
        box(ws, 10, col, f"=1/(1+Premissas!$B$4)^{y}", font=F_FORM, num=NUM)          # fator de desconto
        box(ws, 11, col, f"={L}9*{L}10", font=F_FORM, num=CUR_MM)                     # VP
    box(ws, 13, 1, "rNPV (R$ mm)", font=F_HEAD)
    box(ws, 13, 2, "=SUM(B11:M11)", font=F_FORM, num=CUR_MM)
    widths(ws, [34] + [12] * 13)

    # ---------- 4. Cenários ----------
    ws = wb.create_sheet("Cenarios")
    title(ws, "CENARIOS — First Chicago (prob. somam 100%)", 5)
    header(ws, 3, ["Cenario", "Probabilidade", "Valor (R$ mm)", "Premissa-chave", "Obs"])
    cen = [("Downside", 0.25, None, "", ""), ("Base", 0.50, None, "", ""), ("Case", 0.25, None, "", "")]
    r = 4
    for c in cen:
        box(ws, r, 1, c[0], font=F_HEAD)
        box(ws, r, 2, c[1], font=F_INPUT, num=PCT)
        box(ws, r, 3, c[2], font=F_FORM, num=CUR_MM)
        r += 1
    box(ws, r, 1, "Checagem (deve dar 100%)", font=F_HEAD)
    box(ws, r, 2, "=SUM(B4:B6)", font=F_FORM, num=PCT)
    box(ws, r + 1, 1, "Valor ponderado (R$ mm)", font=F_HEAD)
    box(ws, r + 1, 2, "=SUMPRODUCT($B$4:$B$6,C4:C6)", font=F_FORM, num=CUR_MM)
    widths(ws, [30, 16, 16, 40, 24])

    # ---------- 5. Comparables ----------
    ws = wb.create_sheet("Comparables")
    title(ws, "COMPARABLES — deals de licenciamento + rounds", 7)
    header(ws, 3, ["Tipo", "Empresa/Deal", "Upfront (R$ mm)", "Milestones (R$ mm)", "Royalty", "Data", "Confianca"])
    comp = [
        ("Deal", "Ex.: Prime x BMS", 110, 3500, "a definir", "2024-09", "ALTA"),
        ("Round", "Ex.: Beam Series A", None, None, "", "2018", "MEDIA"),
    ]
    r = 4
    for c in comp:
        for j, v in enumerate(c, start=1):
            num = CUR_MM if j in (3, 4) else PCT if j == 5 and isinstance(v, float) else None
            box(ws, r, j, v, font=F_INPUT if j <= 5 else F_FORM, num=num)
        r += 1
    widths(ws, [14, 30, 18, 18, 12, 12, 12])

    # ---------- 6. Reverse ----------
    ws = wb.create_sheet("Reverse")
    title(ws, "REVERSE — o que precisa ser verdade", 5)
    header(ws, 3, ["Item", "Valor", "Base rate / par", "Plausivel?", "Obs"])
    rev = [
        ("Pico implícito p/ sustentar o valor (R$ mm)", None, None, "", "derivar do Modelo"),
        ("POS implícita vs tabela publica", None, None, "", "comparar com Wong 2019"),
    ]
    r = 4
    for v in rev:
        for j, x in enumerate(v, start=1):
            box(ws, r, j, x, font=F_INPUT if j == 2 else F_FORM, num=CUR_MM if j == 2 else None)
        r += 1
    widths(ws, [46, 16, 18, 12, 34])

    # ---------- 7. CapTable ----------
    ws = wb.create_sheet("CapTable")
    title(ws, "CAP TABLE — diluição e múltiplo do investidor", 6)
    header(ws, 3, ["Item", "Valor", "Formula / Obs", "", "", ""])
    cap = [
        ("Pre-money (R$ mm)", None, "input"),
        ("Investimento (R$ mm)", None, "input"),
        ("Post-money (R$ mm)", "=B4+B5", "= pre + investimento"),
        ("% investidor", "=B5/B6", "= investimento / post"),
        ("Option pool pos-rodada", 0.10, "input"),
        ("Valor de saida cenario base (R$ mm)", None, "input"),
        ("Multiplo do investidor (x)", "=B9*B7/B5", "= saida x % / investimento"),
    ]
    r = 4
    for c in cap:
        box(ws, r, 1, c[0], font=F_HEAD)
        num = PCT if isinstance(c[1], float) else CUR_MM if r in (4, 5, 6, 9) else "0.0x" if r == 10 else None
        box(ws, r, 2, c[1], font=F_INPUT if r in (4, 5, 8, 9) else F_FORM, num=num)
        box(ws, r, 3, c[2], font=F_FORM)
        r += 1
    widths(ws, [34, 18, 30])

    # ---------- 8. Fontes ----------
    ws = wb.create_sheet("Fontes")
    title(ws, "FONTES — registro de premissas", 7)
    header(ws, 3, ["Premissa", "Valor", "Fonte", "URL/DOI", "Data", "Confianca", "Status"])
    widths(ws, [34, 14, 30, 40, 12, 12, 14])

    # ---------- 9. Resumo ----------
    ws = wb.create_sheet("Resumo")
    title(ws, "RESUMO — football field, posicionamento e narrativa", 6)
    header(ws, 3, ["Metodo", "Pior (R$ mm)", "Melhor (R$ mm)", "Central (R$ mm)", "Peso", "Obs"])
    ff = [("rNPV", None, None, None, "", ""), ("Deals", None, None, None, "", ""),
          ("Rounds", None, None, None, "", ""), ("Cenarios", None, None, None, "", "")]
    r = 4
    for f in ff:
        for j, v in enumerate(f, start=1):
            box(ws, r, j, v, font=F_INPUT if j <= 4 else F_FORM, num=CUR_MM if j in (2, 3, 4) else None)
        r += 1
    box(ws, r, 1, "Posicionamento da rodada (R$ mm)", font=F_HEAD)
    box(ws, r, 2, None, font=F_INPUT, num=CUR_MM)
    box(ws, r, 3, "Limite inferior do meio da faixa", font=F_FORM)
    box(ws, r + 2, 1, "Narrativa (3 frases p/ investidor)", font=F_HEAD)
    box(ws, r + 3, 1, None, font=F_FORM)
    widths(ws, [34, 16, 16, 16, 10, 36])

    wb.save(path)
    print(f"Workbook criado: {path}")
    print(f"Abas: {wb.sheetnames}")


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "Valuation_Template_v1.xlsx"
    build(out)
